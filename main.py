import pandas as pd
import warnings
from statsmodels.tsa.stattools import adfuller
import pmdarima as pm
from statsmodels.tsa.statespace.sarimax import SARIMAX # модель Sarimax
from openpyxl import load_workbook # Эксель
import time # время


def forecast(name, arimaParamets, sarimaParamets):
    timeSeries = pd.read_excel(f'../CryptoFile/Postcoin/Post{name}', usecols='B:C')

    timeSeries = timeSeries.fillna(timeSeries.mean())
    timeSeries['Дата'] = pd.to_datetime(timeSeries['Дата'])
    timeSeries.set_index('Дата', inplace=True)


    print('Результат теста:')
    dfResult = adfuller(timeSeries['Цена'])
    dfLabels = ['ADF Test Statistic', 'p-value', '#Lags Used', 'Number of Observations Used']
    for resultValue, label in zip(dfResult, dfLabels):
        print(label + ' : ' + str(resultValue))

    if dfResult[1] <= 0.05:
        print("Сильные доказательства против нулевой гипотезы, ряд является стационарным.")
    else:
        print("Слабые доказательства против нулевой гипотезы, ряд не является стационарным.")



    warnings.filterwarnings("ignore")

    model = SARIMAX(timeSeries, order=arimaParamets, seasonal_order=sarimaParamets)
    results = model.fit()
    print(results.summary())


    stPred = results.get_prediction(start=pd.to_datetime('2023-05-11'), dynamic=False) # Начальная точка прогноза
    forecastValues = stPred.predicted_mean


    actualValues = timeSeries['2023-05-11':]['Цена']
    forecastMse = ((forecastValues - actualValues) ** 2).mean()
    print('Среднеквадратичная ошибка прогноза составляет {}'.format(round(forecastMse, 2)))


    predFuture = results.get_forecast(steps=10) # Прогноз на 10 дней

    print(f'Средние прогнозируемые значения:\n\n{predFuture.predicted_mean}')


    editFile(name, predFuture.predicted_mean)

    print('\n\n')


def testArima(name): # Тест на подбор параметров Sarima
    time_series = pd.read_excel(f'../CryptoFile/Postcoin/Post{name}', usecols='B:C')
    test = pm.auto_arima(time_series, m=7, start_p=0, d=1, start_q=0, start_P=0, D=1, start_Q=0, max_p=3, max_q=1, max_P=3, max_Q=1, trace=True, seasonal=True)
    print(test)


def editFile(name, data): # Редактирует файл в нужный для платформы формат
    file = f'../CryptoFile/Forecastcoin/Forecast{name}'
    wb = load_workbook(file)
    ws = wb['Sheet1']

    i = 2
    for date, price in data.items():
        priceF = float(price)

        ws[f'B{i}'] = date
        ws[f'C{i}'] = priceF

        i += 1

    wb.save(file)
    wb.close()



while True:
    forecast('Bitcoin.xlsx', (0,1,0), (0, 1, 1, 7))
    forecast('Ethereum.xlsx', (0,1,0), (2, 1, 1, 7))
    forecast('Litecoin.xlsx', (1,1,0), (3, 1, 1, 7))
    forecast('Dash.xlsx', (0,1,0), (0, 1, 1, 7))
    forecast('Neo.xlsx', (0,1,1), (0, 1, 1, 7))
    forecast('Xrp.xlsx', (1,1,0), (3, 1, 0, 7))

    print('#'* 60)
    time.sleep(5 * 86400) # Раз в пять дней